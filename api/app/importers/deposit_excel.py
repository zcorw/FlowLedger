from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any, Dict, List, Optional

from pathlib import Path

from fastapi import HTTPException, UploadFile
from openpyxl import load_workbook
from pydantic import ValidationError

from ..schemas.deposit_import import (
    ImportBalanceItem,
    ImportDepositRequest,
    ImportInstitutionItem,
    ImportProductItem,
)


def _normalize_str(value: Any) -> Optional[str]:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _normalize_decimal(value: Any) -> Optional[Decimal]:
    if value is None or value == "":
        return None
    if isinstance(value, Decimal):
        return value
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError):
        return None


def _normalize_datetime(value: Any) -> Optional[datetime]:
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day)
    if value is None or value == "":
        return None
    if isinstance(value, str):
        try:
            return datetime.fromisoformat(value)
        except ValueError:
            return None
    return None


def _normalize_enum(value: Any) -> Optional[str]:
    text = _normalize_str(value)
    return text.lower() if text else None


def _normalize_currency(value: Any) -> Optional[str]:
    text = _normalize_str(value)
    return text.upper() if text else None


def _sanitize_sheet_name(name: str) -> str:
    invalid = {":", "\\", "/", "?", "*", "[", "]"}
    cleaned = "".join(ch for ch in name if ch not in invalid).strip()
    if not cleaned:
        cleaned = "Sheet"
    return cleaned[:31]


def _read_sheet_rows(ws, headers: List[str]) -> List[Dict[str, Any]]:
    header_row = [cell.value for cell in ws[1]]
    header_map = {str(h).strip(): idx for idx, h in enumerate(header_row) if h is not None}
    missing = [h for h in headers if h not in header_map]
    if missing:
        missing_list = ",".join(missing)
        raise HTTPException(status_code=422, detail=f"missing_headers:{missing_list}")

    rows: List[Dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(cell is None or str(cell).strip() == "" for cell in row):
            continue
        item: Dict[str, Any] = {}
        for key in headers:
            idx = header_map[key]
            item[key] = row[idx] if idx < len(row) else None
        rows.append(item)
    return rows


def _parse_deposit_import_content(filename: str, content: bytes) -> ImportDepositRequest:
    if not filename or not filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=422, detail="invalid_file_type")

    try:
        wb = load_workbook(BytesIO(content), data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=422, detail="invalid_excel") from exc

    if "Institutions" not in wb.sheetnames:
        raise HTTPException(status_code=422, detail="missing_sheet:Institutions")
    if "Products" not in wb.sheetnames:
        raise HTTPException(status_code=422, detail="missing_sheet:Products")

    inst_rows = _read_sheet_rows(
        wb["Institutions"],
        ["Name", "Type", "Status"],
    )
    prod_rows = _read_sheet_rows(
        wb["Products"],
        ["Name", "Institution", "Type", "Status", "Currency", "Risk Level"],
    )

    institutions: List[ImportInstitutionItem] = []
    for row in inst_rows:
        institutions.append(
            ImportInstitutionItem(
                name=_normalize_str(row["Name"]),
                type=_normalize_enum(row["Type"]),
                status=_normalize_enum(row["Status"]),
            )
        )

    products: List[ImportProductItem] = []
    for row in prod_rows:
        products.append(
            ImportProductItem(
                institution_name=_normalize_str(row["Institution"]),
                name=_normalize_str(row["Name"]),
                product_type=_normalize_enum(row["Type"]) or "deposit",
                currency=_normalize_currency(row["Currency"]),
                status=_normalize_enum(row["Status"]) or "active",
                risk_level=_normalize_enum(row["Risk Level"]) or "stable",
            )
        )

    balances: List[ImportBalanceItem] = []
    institution_names = [item.name for item in institutions]
    for inst_name in institution_names:
        sheet_name = inst_name
        if sheet_name not in wb.sheetnames:
            sheet_name = _sanitize_sheet_name(inst_name)
        if sheet_name not in wb.sheetnames:
            raise HTTPException(status_code=422, detail=f"missing_sheet:{inst_name}")

        ws = wb[sheet_name]
        header_row = [cell.value for cell in ws[1]]
        if not header_row or str(header_row[0]).strip().lower() != "date":
            raise HTTPException(status_code=422, detail=f"invalid_balance_header:{inst_name}")
        product_headers = []
        for cell in header_row[1:]:
            name = _normalize_str(cell)
            product_headers.append(name)

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or (row[0] is None and all(cell is None for cell in row[1:])):
                continue
            as_of = _normalize_datetime(row[0])
            if not as_of:
                continue
            for idx, amount in enumerate(row[1:], start=0):
                product_name = product_headers[idx] if idx < len(product_headers) else None
                if not product_name:
                    continue
                value = _normalize_decimal(amount)
                if value is None:
                    continue
                balances.append(
                    ImportBalanceItem(
                        institution_name=inst_name,
                        product_name=product_name,
                        as_of=as_of,
                        amount=value,
                    )
                )

    try:
        return ImportDepositRequest(
            institutions=institutions,
            products=products,
            product_balances=balances,
        )
    except ValidationError as exc:
        raise HTTPException(status_code=422, detail=exc.errors()) from exc


def parse_deposit_import_file(upload: UploadFile) -> ImportDepositRequest:
    content = upload.file.read()
    return _parse_deposit_import_content(upload.filename or "", content)


def parse_deposit_import_path(path: Path) -> ImportDepositRequest:
    content = path.read_bytes()
    return _parse_deposit_import_content(path.name, content)
