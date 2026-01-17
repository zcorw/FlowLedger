from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any, Dict, List, Optional

from fastapi import HTTPException, UploadFile
from openpyxl import load_workbook
from pydantic import ValidationError

from ..schemas.exchange_rate_import import (
    ImportExchangeRateItem,
    ImportExchangeRateRequest,
)


def _normalize_str(value: Any) -> Optional[str]:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _normalize_decimal(value: Any) -> Optional[Decimal]:
    if value is None or value == "":
        return None
    if isinstance(value, Decimal):
        return value
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError):
        return None


def _normalize_date(value: Any) -> Optional[date]:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if value is None or value == "":
        return None
    try:
        return date.fromisoformat(str(value).strip())
    except ValueError:
        return None


def _read_sheet_rows(
    ws,
    required_headers: List[str],
    optional_headers: List[str],
) -> List[Dict[str, Any]]:
    header_row = [cell.value for cell in ws[1]]
    header_map = {str(h).strip(): idx for idx, h in enumerate(header_row) if h is not None}
    missing = [h for h in required_headers if h not in header_map]
    if missing:
        missing_list = ",".join(missing)
        raise HTTPException(status_code=422, detail=f"missing_headers:{missing_list}")

    keys = required_headers + [h for h in optional_headers if h in header_map]
    rows: List[Dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(cell is None or str(cell).strip() == "" for cell in row):
            continue
        item: Dict[str, Any] = {}
        for key in keys:
            idx = header_map[key]
            item[key] = row[idx] if idx < len(row) else None
        rows.append(item)
    return rows


def parse_exchange_rate_import_file(upload: UploadFile) -> ImportExchangeRateRequest:
    if not upload.filename or not upload.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=422, detail="invalid_file_type")

    content = upload.file.read()
    try:
        wb = load_workbook(BytesIO(content), data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=422, detail="invalid_excel") from exc

    if "exchange_rates" not in wb.sheetnames:
        raise HTTPException(status_code=422, detail="missing_sheet:exchange_rates")

    rows = _read_sheet_rows(
        wb["exchange_rates"],
        ["base", "quote", "rate_date", "rate"],
        ["source"],
    )

    items: List[ImportExchangeRateItem] = []
    for row in rows:
        items.append(
            ImportExchangeRateItem(
                base=_normalize_str(row["base"]),
                quote=_normalize_str(row["quote"]),
                rate_date=_normalize_date(row["rate_date"]),
                rate=_normalize_decimal(row["rate"]),
                source=_normalize_str(row.get("source")),
            )
        )

    try:
        return ImportExchangeRateRequest(items=items)
    except ValidationError as exc:
        raise HTTPException(status_code=422, detail=exc.errors()) from exc
