from __future__ import annotations

from dataclasses import dataclass
from typing import Dict, Iterable

from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

_INVALID_SHEET_CHARS = {":", "\\", "/", "?", "*", "[", "]"}
_MAX_SHEET_NAME_LEN = 31


@dataclass
class ExcelWorkbookContext:
    workbook: Workbook
    sheets: Dict[str, Worksheet]
       

def _sanitize_sheet_name(name: str) -> str:
    cleaned = "".join(ch for ch in name if ch not in _INVALID_SHEET_CHARS).strip()
    if not cleaned:
        cleaned = "Sheet"
    return cleaned[:_MAX_SHEET_NAME_LEN]


def _dedupe_sheet_name(base: str, used: set[str]) -> str:
    if base not in used:
        return base
    counter = 2
    while True:
        suffix = f"_{counter}"
        trimmed = base[: _MAX_SHEET_NAME_LEN - len(suffix)]
        candidate = f"{trimmed}{suffix}"
        if candidate not in used:
            return candidate
        counter += 1


def create_workbook_with_sheets(sheet_names: Iterable[str]) -> ExcelWorkbookContext:
    wb = Workbook()
    names = list(sheet_names)
    sheets: Dict[str, Worksheet] = {}
    if not names:
        ws = wb.active
        sheets[ws.title] = ws
        return ExcelWorkbookContext(workbook=wb, sheets=sheets)

    wb.remove(wb.active)
    used: set[str] = set()
    for raw_name in names:
        base = _sanitize_sheet_name(str(raw_name))
        final_name = _dedupe_sheet_name(base, used)
        ws = wb.create_sheet(title=final_name)
        sheets[final_name] = ws
        used.add(final_name)

    return ExcelWorkbookContext(workbook=wb, sheets=sheets)

_CURRENCY_DISPLAY_FORMATS: Dict[str, str] = {
    "CNY": "[$CNY]#,##0.00",
    "USD": "[$USD]#,##0.00",
    "HKD": "[$HKD]#,##0.00",
    "JPY": "[$JPY]#,##0",
}

def _get_currency_display_format(currency: str) -> str:
    code = (currency or "").upper().strip()
    return _CURRENCY_DISPLAY_FORMATS.get(code, "[$]#,##0.00")


def apply_currency_display_format(cell, currency: str) -> None:
    cell.number_format = _get_currency_display_format(currency)

def freeze_panes(ws: Worksheet, *, cell_range: str = "A2") -> None:
    ws.freeze_panes = ws[cell_range]


def add_dropdown_list(
    ws: Worksheet,
    cell_range: str,
    options: Iterable[str] | str,
    *,
    allow_blank: bool = True,
) -> DataValidation:
    if isinstance(options, str):
        formula1 = options
        if not (formula1.startswith("=") or (formula1.startswith('"') and formula1.endswith('"'))):
            formula1 = f'"{formula1}"'
    else:
        items = [str(item) for item in options]
        formula1 = f'"{",".join(items)}"'
    # showDropDown=False means Excel will show the dropdown arrow.
    dv = DataValidation(type="list", formula1=formula1, allow_blank=allow_blank, showDropDown=False)
    ws.add_data_validation(dv)
    dv.add(cell_range)
    return dv
