from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal
from io import BytesIO
from typing import Dict, List

from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from ..excel_utils import (
    add_dropdown_list,
    apply_currency_display_format,
    create_workbook_with_sheets,
    freeze_panes,
)


@dataclass
class ProductExcelData:
    id: int
    name: str
    product_type: str
    currency: str
    status: str
    risk_level: str


@dataclass
class InstitutionExcelData:
    id: int
    name: str
    type: str
    status: str
    products: List[ProductExcelData]
    balances: List[ProductBalanceExcelData]


@dataclass
class ProductBalanceExcelData:
    product_id: int
    as_of: datetime
    balance: Decimal


@dataclass
class ExportExcelData:
    institutions: List[InstitutionExcelData]


@dataclass
class ExportBalanceItem:
    product_name: str
    currency: str
    as_of: datetime
    balance: Decimal


def _institution_to_sheet(inst: List[InstitutionExcelData], ws: Worksheet) -> None:
    add_dropdown_list(ws, "B2:B100", ["BANK", "BROKER", "OTHER"])
    add_dropdown_list(ws, "C2:C100", ["ACTIVE", "INACTIVE"])

    sheet_title = ["Name", "Type", "Status"]
    ws.append(sheet_title)
    for item in inst:
        row = [item.name, item.type.upper(), item.status.upper()]
        ws.append(row)
    freeze_panes(ws)


def _product_to_sheet(prod: Dict[str, List[ProductExcelData]], ws: Worksheet) -> None:
    add_dropdown_list(ws, "C2:C100", ["DEPOSIT", "INVESTMENT", "SECURITIES", "OTHER"])
    add_dropdown_list(ws, "D2:D100", ["ACTIVE", "INACTIVE"])
    add_dropdown_list(ws, "F2:F100", ["FLEXIBLE", "STABLE", "HIGH_RISK"])

    sheet_title = ["Name", "Institution", "Type", "Status", "Currency", "Risk Level"]
    ws.append(sheet_title)
    for institution_name, products in prod.items():
        for item in products:
            row = [
                item.name,
                institution_name,
                item.product_type.upper(),
                item.status.upper(),
                item.currency.upper(),
                item.risk_level.upper(),
            ]
            ws.append(row)
    freeze_panes(ws)


def _balance_to_sheet(bal: List[ExportBalanceItem], products: List[str], ws: Worksheet) -> None:
    sheet_title = ["Date"] + products
    product_col = {
        product_name: get_column_letter(idx + 2) for idx, product_name in enumerate(products)
    }
    ws.append(sheet_title)
    row_num = ws.max_row
    current_date = None
    # Sort by as_of ascending.
    for item in sorted(bal, key=lambda x: x.as_of):
        date_str = item.as_of.strftime("%Y-%m-%d")
        if date_str != current_date:
            ws.append([date_str] + [None] * len(products))
            current_date = date_str
            row_num = ws.max_row
        col = product_col.get(item.product_name)
        if not col:
            continue
        current_cell = ws[f"{col}{row_num}"]
        apply_currency_display_format(current_cell, item.currency)
        current_cell.value = item.balance
    freeze_panes(ws)


def export_to_excel(data: ExportExcelData) -> bytes:
    sheet_names = ["Institutions", "Products"] + [item.name for item in data.institutions]
    cx = create_workbook_with_sheets(sheet_names)
    _institution_to_sheet(data.institutions, cx.sheets["Institutions"])
    inst_products = {item.name: item.products for item in data.institutions}
    _product_to_sheet(inst_products, cx.sheets["Products"])

    inst_sheet_name_map: Dict[str, str] = {}
    ordered_sheet_names = cx.workbook.sheetnames
    for idx, inst in enumerate(data.institutions):
        inst_sheet_name_map[inst.name] = ordered_sheet_names[idx + 2]

    for item in data.institutions:
        inst_product = inst_products[item.name]
        products_name = [product.name for product in inst_product]
        balances: List[ExportBalanceItem] = []
        for balance in item.balances:
            product = next((p for p in inst_product if p.id == balance.product_id), None)
            if product:
                balances.append(
                    ExportBalanceItem(
                        product.name,
                        product.currency,
                        balance.as_of,
                        balance.balance,
                    )
                )
        _balance_to_sheet(balances, products_name, cx.sheets[inst_sheet_name_map[item.name]])

    output = BytesIO()
    cx.workbook.save(output)
    return output.getvalue()
